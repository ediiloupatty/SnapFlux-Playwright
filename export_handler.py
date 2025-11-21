"""
Export handler untuk export hasil automation ke Excel dengan format multi-tanggal
File ini menangani export data ke Excel dengan kolom dinamis per tanggal
"""

import os
from datetime import datetime
from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def export_results_to_excel(
    results: List[Dict], export_date: datetime, custom_filepath: str = None
) -> str:
    """
    Export hasil automation ke Excel dengan format multi-tanggal

    Args:
        results (List[Dict]): List hasil automation
        export_date (datetime): Tanggal export
        custom_filepath (str, optional): Path custom untuk menyimpan file. Defaults to None.

    Returns:
        str: Path file Excel yang di-generate
    """
    try:
        if custom_filepath:
            filepath = custom_filepath
            # Ensure directory exists
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
        else:
            # Buat folder results jika belum ada
            results_dir = os.path.join(os.path.dirname(__file__), "results")
            os.makedirs(results_dir, exist_ok=True)

            # Nama file dengan timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"SnapFlux_Export_{timestamp}.xlsx"
            filepath = os.path.join(results_dir, filename)

        # Buat workbook baru
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export Data"

        # Format tanggal untuk header
        date_str = export_date.strftime("%Y-%m-%d")

        # HEADER ROW 1: PANGKALAN_ID, NAMA_PANGKALAN, dan tanggal-tanggal
        ws.cell(row=1, column=1, value="PANGKALAN_ID")
        ws.cell(row=1, column=2, value="NAMA_PANGKALAN")
        ws.cell(row=1, column=3, value=date_str)

        # HEADER ROW 2: Sub-headers untuk setiap tanggal
        ws.cell(row=2, column=1, value="")  # Kosong untuk PANGKALAN_ID
        ws.cell(row=2, column=2, value="")  # Kosong untuk NAMA_PANGKALAN
        ws.cell(row=2, column=3, value="STOK (TABUNG)")
        ws.cell(row=2, column=4, value="INPUT (TABUNG)")
        ws.cell(row=2, column=5, value="STATUS")

        # Merge cells untuk header tanggal
        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        # Style untuk header
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Apply style ke header
        for col in range(1, 6):
            cell1 = ws.cell(row=1, column=col)
            cell2 = ws.cell(row=2, column=col)
            cell1.fill = header_fill
            cell1.font = header_font
            cell1.alignment = header_alignment
            cell1.border = border
            cell2.fill = header_fill
            cell2.font = header_font
            cell2.alignment = header_alignment
            cell2.border = border

        # DATA ROWS
        current_row = 3
        for result in results:
            # Extract data
            pangkalan_id = result.get("pangkalan_id", "")
            nama = result.get("nama", "")
            stok = result.get("stok", "0 Tabung").replace(" Tabung", "")
            tabung_terjual = result.get("tabung_terjual", "0 Tabung").replace(
                " Tabung", ""
            )
            status = result.get("status", "Tidak Ada Penjualan")

            # Convert stok dan tabung_terjual ke int
            try:
                stok_int = int(stok)
            except:
                stok_int = 0

            try:
                tabung_int = int(tabung_terjual)
            except:
                tabung_int = 0

            # Write data
            ws.cell(row=current_row, column=1, value=pangkalan_id)
            ws.cell(row=current_row, column=2, value=nama)
            ws.cell(row=current_row, column=3, value=stok_int)
            ws.cell(row=current_row, column=4, value=tabung_int)
            ws.cell(row=current_row, column=5, value=status)

            # Style untuk data rows
            data_alignment = Alignment(horizontal="left", vertical="center")
            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                if col in [3, 4]:  # Number columns
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = data_alignment

                # Conditional formatting untuk status
                if col == 5:
                    if status == "Ada Penjualan":
                        cell.fill = PatternFill(
                            start_color="C6EFCE",
                            end_color="C6EFCE",
                            fill_type="solid",
                        )
                        cell.font = Font(color="006100")
                    else:
                        cell.fill = PatternFill(
                            start_color="FFC7CE",
                            end_color="FFC7CE",
                            fill_type="solid",
                        )
                        cell.font = Font(color="9C0006")

            current_row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 30  # PANGKALAN_ID
        ws.column_dimensions["B"].width = 35  # NAMA_PANGKALAN
        ws.column_dimensions["C"].width = 15  # STOK
        ws.column_dimensions["D"].width = 15  # INPUT
        ws.column_dimensions["E"].width = 20  # STATUS

        # Freeze panes (freeze header rows)
        ws.freeze_panes = "A3"

        # Save workbook
        wb.save(filepath)

        return filepath

    except Exception as e:
        print(f"✗ Error exporting to Excel: {str(e)}")
        raise


def export_multi_date_results(
    all_results: Dict[str, List[Dict]], dates: List[datetime]
) -> str:
    """
    Export hasil automation multi-tanggal ke Excel dengan kolom dinamis

    Args:
        all_results (Dict[str, List[Dict]]): Dictionary dengan key tanggal (YYYY-MM-DD) dan value list hasil
        dates (List[datetime]): List tanggal yang akan di-export

    Returns:
        str: Path file Excel yang di-generate
    """
    try:
        # Buat folder results jika belum ada
        results_dir = os.path.join(os.path.dirname(__file__), "results")
        os.makedirs(results_dir, exist_ok=True)

        # Nama file dengan timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"SnapFlux_MultiDate_Export_{timestamp}.xlsx"
        filepath = os.path.join(results_dir, filename)

        # Buat workbook baru
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Multi-Date Export"

        # HEADER ROW 1: PANGKALAN_ID, NAMA_PANGKALAN, dan tanggal-tanggal
        ws.cell(row=1, column=1, value="PANGKALAN_ID")
        ws.cell(row=1, column=2, value="NAMA_PANGKALAN")

        # Merge untuk kolom ID dan Nama
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        # Header untuk setiap tanggal (3 kolom per tanggal)
        current_col = 3
        date_columns = {}  # Map tanggal ke kolom awal

        for date in sorted(dates):
            date_str = date.strftime("%Y-%m-%d")
            date_columns[date_str] = current_col

            # Merge 3 kolom untuk tanggal
            ws.merge_cells(
                start_row=1,
                start_column=current_col,
                end_row=1,
                end_column=current_col + 2,
            )
            ws.cell(row=1, column=current_col, value=date_str)

            # Sub-headers
            ws.cell(row=2, column=current_col, value="STOK (TABUNG)")
            ws.cell(row=2, column=current_col + 1, value="INPUT (TABUNG)")
            ws.cell(row=2, column=current_col + 2, value="STATUS")

            current_col += 3

        # Style untuk header
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Apply style ke header
        for col in range(1, current_col):
            cell1 = ws.cell(row=1, column=col)
            cell2 = ws.cell(row=2, column=col)
            cell1.fill = header_fill
            cell1.font = header_font
            cell1.alignment = header_alignment
            cell1.border = border
            cell2.fill = header_fill
            cell2.font = header_font
            cell2.alignment = header_alignment
            cell2.border = border

        # Collect all unique accounts
        all_accounts = {}
        for date_str, results in all_results.items():
            for result in results:
                pangkalan_id = result.get("pangkalan_id", "")
                if pangkalan_id not in all_accounts:
                    all_accounts[pangkalan_id] = {
                        "nama": result.get("nama", ""),
                        "data": {},
                    }

                all_accounts[pangkalan_id]["data"][date_str] = result

        # DATA ROWS
        current_row = 3
        for pangkalan_id, account_data in sorted(all_accounts.items()):
            # Write pangkalan ID and nama
            ws.cell(row=current_row, column=1, value=pangkalan_id)
            ws.cell(row=current_row, column=2, value=account_data["nama"])

            # Write data untuk setiap tanggal
            for date in sorted(dates):
                date_str = date.strftime("%Y-%m-%d")
                start_col = date_columns[date_str]

                if date_str in account_data["data"]:
                    result = account_data["data"][date_str]

                    # Extract data
                    stok = result.get("stok", "0 Tabung").replace(" Tabung", "")
                    tabung_terjual = result.get("tabung_terjual", "0 Tabung").replace(
                        " Tabung", ""
                    )
                    status = result.get("status", "Tidak Ada Penjualan")

                    # Convert to int
                    try:
                        stok_int = int(stok)
                    except:
                        stok_int = 0

                    try:
                        tabung_int = int(tabung_terjual)
                    except:
                        tabung_int = 0

                    ws.cell(row=current_row, column=start_col, value=stok_int)
                    ws.cell(row=current_row, column=start_col + 1, value=tabung_int)
                    ws.cell(row=current_row, column=start_col + 2, value=status)
                else:
                    # No data untuk tanggal ini
                    ws.cell(row=current_row, column=start_col, value=0)
                    ws.cell(row=current_row, column=start_col + 1, value=0)
                    ws.cell(
                        row=current_row,
                        column=start_col + 2,
                        value="Tidak Ada Data",
                    )

            # Apply style to data row
            for col in range(1, current_col):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border

                if col <= 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif (col - 3) % 3 in [0, 1]:  # STOK and INPUT columns
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:  # STATUS column
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    status_value = cell.value or ""
                    if "Ada Penjualan" in status_value:
                        cell.fill = PatternFill(
                            start_color="C6EFCE",
                            end_color="C6EFCE",
                            fill_type="solid",
                        )
                        cell.font = Font(color="006100")
                    elif "Tidak Ada Penjualan" in status_value:
                        cell.fill = PatternFill(
                            start_color="FFC7CE",
                            end_color="FFC7CE",
                            fill_type="solid",
                        )
                        cell.font = Font(color="9C0006")
                    else:
                        cell.fill = PatternFill(
                            start_color="FFEB9C",
                            end_color="FFEB9C",
                            fill_type="solid",
                        )
                        cell.font = Font(color="9C5700")

            current_row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 30  # PANGKALAN_ID
        ws.column_dimensions["B"].width = 35  # NAMA_PANGKALAN

        for col in range(3, current_col):
            col_letter = get_column_letter(col)
            if (col - 3) % 3 in [0, 1]:  # STOK and INPUT
                ws.column_dimensions[col_letter].width = 15
            else:  # STATUS
                ws.column_dimensions[col_letter].width = 20

        # Freeze panes
        ws.freeze_panes = "C3"

        # Save workbook
        wb.save(filepath)

        return filepath

    except Exception as e:
        print(f"✗ Error exporting multi-date to Excel: {str(e)}")
        raise
