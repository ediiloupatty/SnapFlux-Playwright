"""
Excel handler untuk mengelola operasi Excel
File ini berisi semua fungsi untuk menangani Excel operations
"""

import os
import re
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from modules.core.constants import (
        BULAN_ID,
        RESULTS_DIR,
        get_master_filename,
        get_sheet_name_dynamic,
    )
    from modules.core.validators import parse_inputan_to_int, parse_stok_to_int
except ImportError:
    from modules.core.constants import (
        BULAN_ID,
        RESULTS_DIR,
        get_master_filename,
        get_sheet_name_dynamic,
    )
    from modules.core.validators import parse_inputan_to_int, parse_stok_to_int


def get_excel_filename(selected_date=None):
    """Generate nama file Excel sesuai format"""
    if selected_date:
        tgl = selected_date.day
        bln = BULAN_ID[selected_date.month]
        thn = selected_date.year
        return f"DATA TRANSAKSI SNAPFLUX PANGKALAN {tgl} {bln} {thn}.xlsx"
    else:
        now = datetime.now()
        return f"DATA TRANSAKSI SNAPFLUX PANGKALAN TANPA FILTER TANGGAL {now.strftime('%d %B %Y')}.xlsx"


def get_day_color(date_str):
    """Return fill color berdasarkan hari dalam seminggu"""
    try:
        # Parse tanggal dari format YYYY-MM-DD
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        weekday = date_obj.weekday()  # 0=Monday, 1=Tuesday, ..., 6=Sunday
        # Default warna putih untuk semua hari
        return "FFFFFF"
    except:
        return "FFFFFF"  # Putih jika error parsing tanggal


def save_to_excel_pivot_format(
    pangkalan_id,
    nama_pangkalan,
    tanggal_check,
    stok_awal,
    total_inputan,
    status,
    selected_date=None,
):
    """
    ============================================
    FUNGSI SIMPAN DATA KE EXCEL DENGAN FORMAT PIVOT
    ============================================

    Fungsi ini menyimpan data hasil scraping ke file Excel dengan format pivot
    yang optimal untuk analisis data. Menggunakan sistem incremental untuk
    menambahkan data baru ke file master yang sudah ada.

    Proses yang dilakukan:
    1. Generate nama file Excel berdasarkan tanggal (master format)
    2. Parse data stok dan inputan dari string ke integer
    3. Tentukan tanggal operasi (dari filter atau tanggal sekarang)
    4. Generate timestamp dengan format AM/PM
    5. Cek apakah file Excel sudah ada:
       - Jika ada: Load file dan tambahkan data baru
       - Jika tidak ada: Buat file baru dengan header
    6. Tambahkan data baru ke sheet yang sesuai
    7. Apply formatting (border, alignment, warna hari)
    8. Simpan file Excel

    Format data yang disimpan:
    - PANGKALAN_ID: ID pangkalan merchant
    - NAMA_PANGKALAN: Nama pangkalan merchant
    - TANGGAL_CHECK: Tanggal operasi
    - STOK_AWAL: Stok awal dari dashboard
    - TOTAL_INPUTAN: Total tabung terjual
    - STATUS: Status penjualan (Ada Penjualan/Tidak Ada Penjualan)
    - TIME: Timestamp operasi

    Args:
        pangkalan_id (str): ID pangkalan merchant
        nama_pangkalan (str): Nama pangkalan merchant
        tanggal_check (str): Tanggal check dalam format string
        stok_awal (str): Stok awal dari dashboard
        total_inputan (str): Total inputan dalam format "X Tabung"
        status (str): Status penjualan
        selected_date (datetime): Tanggal yang dipilih user (optional)

    Returns:
        None: Menyimpan data ke file Excel
    """

    # Gunakan file master untuk sistem incremental
    filename = get_master_filename(selected_date)
    filepath = os.path.join(RESULTS_DIR, filename)

    try:
        # Parse data
        stok_int = parse_stok_to_int(stok_awal)
        inputan_int = parse_inputan_to_int(total_inputan)

        # Tentukan tanggal untuk operasi
        if selected_date:
            target_date = selected_date
        else:
            target_date = datetime.now()

        # Timestamp untuk TIME dengan format AM/PM
        timestamp = datetime.now().strftime("%I:%M %p")

        print(
            f"Parsing data: stok='{stok_awal}' -> {stok_int}, inputan='{total_inputan}' -> {inputan_int}"
        )

        # Generate sheet name dinamis berdasarkan bulan
        sheet_name = get_sheet_name_dynamic(selected_date)

        print(f"Menggunakan file master: {filename}")
        print(f"Menggunakan sheet: {sheet_name}")

        # Load existing Excel atau buat baru
        wb, ws = _load_or_create_workbook(filepath, sheet_name)

        # **SISTEM INCREMENTAL BARU** - Cari atau buat kolom untuk tanggal ini
        date_col_start, is_new_date_column = _find_or_create_date_column(
            ws, target_date
        )

        # **SISTEM INCREMENTAL BARU** - Cari atau buat row untuk pangkalan ini
        pangkalan_row, is_new_pangkalan = _find_or_create_pangkalan_row(
            ws, pangkalan_id, nama_pangkalan
        )

        # Update data langsung ke posisi yang tepat dengan center alignment
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Set data dengan center alignment untuk STOK, INPUT, TIME, dan STATUS
        ws.cell(
            row=pangkalan_row, column=date_col_start, value=stok_int
        ).alignment = center_alignment
        ws.cell(
            row=pangkalan_row, column=date_col_start + 1, value=inputan_int
        ).alignment = center_alignment
        ws.cell(
            row=pangkalan_row, column=date_col_start + 2, value=timestamp
        ).alignment = center_alignment
        ws.cell(
            row=pangkalan_row, column=date_col_start + 3, value=status
        ).alignment = center_alignment

        # Apply conditional formatting untuk stok > 90 dan input = 0 (warna kuning)
        _apply_conditional_formatting_new(
            ws, stok_int, inputan_int, pangkalan_row, date_col_start
        )

        # Format headers dengan merge cell dan center alignment dengan fill color berdasarkan hari
        _format_headers_new(ws, target_date.strftime("%Y-%m-%d"))

        # Apply border untuk semua cell
        _apply_borders(ws)

        # Save workbook
        wb.save(filepath)

        print(f"✓ Data berhasil disimpan ke file master incremental: {filepath}")
        print(f"Format: PANGKALAN_ID={pangkalan_id}, NAMA={nama_pangkalan}")
        print(f"Data: STOK={stok_int}, INPUT={inputan_int}, TIME={timestamp}")
        print(f"Sistem incremental aktif - data ditambahkan/updated tanpa overwrite")

    except Exception as e:
        print(f"✗ Error saat menyimpan pivot format: {str(e)}")
        import logging

        logger = logging.getLogger("automation")
        logger.error(f"Error saving to pivot Excel: {str(e)}", exc_info=True)


def _load_or_create_workbook(filepath, sheet_name):
    """Load existing workbook atau buat yang baru dengan sheet name dinamis"""
    if os.path.exists(filepath):
        try:
            wb = load_workbook(filepath)
            # Cek apakah sheet dengan nama dinamis ada
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
            else:
                ws = wb[sheet_name]

        except Exception as e:
            print(f"⚠ Error membaca file, akan buat file baru: {e}")
            # Buat workbook baru
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            ws = wb.create_sheet(sheet_name)
    else:
        # Buat workbook baru
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        ws = wb.create_sheet(sheet_name)

    return wb, ws


def _find_or_create_date_column(ws, target_date):
    """
    Cari kolom untuk tanggal tertentu, buat baru jika belum ada
    Returns: (date_col_start, is_new_column)
    """
    display_date = target_date.strftime("%Y-%m-%d")

    # Cari kolom apakah sudah ada
    for col in range(
        3, ws.max_column + 1
    ):  # Mulai dari kolom 3 (setelah PANGKALAN_ID dan NAMA_PANGKALAN)
        header_cell = ws.cell(row=1, column=col).value
        if header_cell and str(header_cell) == display_date:
            print(f"Kolom tanggal {display_date} sudah ada di kolom {col}")
            return col, False

    # Buat kolom baru jika belum ada
    new_col = ws.max_column + 1
    if ws.max_column < 3:  # File baru, mulai dari kolom 3
        new_col = 3
        # Pastikan header dasar sudah ada untuk file baru
        if (
            ws.cell(row=1, column=1).value is None
            or ws.cell(row=1, column=1).value == ""
        ):
            ws.cell(row=1, column=1, value="PANGKALAN_ID")
            ws.cell(row=1, column=2, value="NAMA_PANGKALAN")

    ws.cell(row=1, column=new_col, value=display_date)
    ws.cell(row=2, column=new_col, value="STOK (TABUNG)")
    ws.cell(row=2, column=new_col + 1, value="INPUT (TABUNG)")
    ws.cell(row=2, column=new_col + 2, value="TIME")
    ws.cell(row=2, column=new_col + 3, value="STATUS")

    print(f"🆕 Membuat kolom baru untuk tanggal {display_date} di kolom {new_col}")
    return new_col, True


def _find_or_create_pangkalan_row(ws, pangkalan_id, nama_pangkalan):
    """
    Cari row untuk pangkalan tertentu, buat baru jika belum ada
    Returns: (row_number, is_new_row)
    """
    # Cari apakah pangkalan sudah ada
    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == pangkalan_id:
            print(f"Pangkalan {pangkalan_id} sudah ada di baris {row}")
            return row, False

    # Buat row baru jika belum ada
    new_row = ws.max_row + 1
    if ws.max_row < 2:  # File baru, mulai dari baris 3
        new_row = 3

    ws.cell(row=new_row, column=1, value=pangkalan_id)
    ws.cell(row=new_row, column=2, value=nama_pangkalan)

    print(f"🆕 Membuat baris baru untuk pangkalan {pangkalan_id} di baris {new_row}")
    return new_row, True


def _setup_headers(ws, display_date, pangkalan_id):
    """Setup headers dan return column start dan row start"""
    date_col_start = 3  # Default untuk file baru
    data_start_row = 3  # Default untuk file baru

    # Cek apakah header sudah ada
    if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
        # Buat header baru sesuai format yang diminta
        # Row 1: PANGKALAN_ID | NAMA_PANGKALAN | ---- DATE ----
        ws.cell(row=1, column=1, value="PANGKALAN_ID")
        ws.cell(row=1, column=2, value="NAMA_PANGKALAN")
        ws.cell(row=1, column=3, value=display_date)

        # Row 2: (kosong) | (kosong) | STOK (TABUNG) | INPUT (TABUNG) | TIME | STATUS
        ws.cell(row=2, column=1, value="")
        ws.cell(row=2, column=2, value="")
        ws.cell(row=2, column=3, value="STOK (TABUNG)")
        ws.cell(row=2, column=4, value="INPUT (TABUNG)")
        ws.cell(row=2, column=5, value="TIME")
        ws.cell(row=2, column=6, value="STATUS")
        return date_col_start, data_start_row
    else:
        # Cek apakah tanggal ini sudah ada di header
        for col in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col).value
            if header_cell and display_date in str(header_cell):
                date_col_start = col
                break

        # Jika tanggal belum ada, tambahkan kolom baru
        if date_col_start == 3 and ws.cell(row=1, column=3).value != display_date:
            date_col_start = ws.max_column + 1
            ws.cell(row=1, column=date_col_start, value=display_date)
            ws.cell(row=2, column=date_col_start, value="STOK (TABUNG)")
            ws.cell(row=2, column=date_col_start + 1, value="INPUT (TABUNG)")
            ws.cell(row=2, column=date_col_start + 2, value="TIME")
            ws.cell(row=2, column=date_col_start + 3, value="STATUS")
        else:
            # Pastikan sub-header tersedia untuk tanggal yang sudah ada
            ws.cell(row=2, column=date_col_start, value="STOK (TABUNG)")
            ws.cell(row=2, column=date_col_start + 1, value="INPUT (TABUNG)")
            ws.cell(row=2, column=date_col_start + 2, value="TIME")
            ws.cell(row=2, column=date_col_start + 3, value="STATUS")

        # Cari row yang tepat untuk data ini
        for row in range(3, ws.max_row + 1):
            pangkal_id_cell = ws.cell(row=row, column=1).value
            if pangkal_id_cell == pangkalan_id:
                data_start_row = row
                break
            elif pangkal_id_cell is None or pangkal_id_cell == "":
                data_start_row = row
                break

        if (
            data_start_row == 3 and ws.max_row > 2
        ):  # Jika tidak ditemukan row yang tepat
            data_start_row = ws.max_row + 1

    return date_col_start, data_start_row


def _fill_data(
    ws,
    pangkalan_id,
    nama_pangkalan,
    stok_int,
    inputan_int,
    timestamp,
    date_col_start,
    data_start_row,
):
    """Fill data ke dalam worksheet"""
    pangkalan_exists = False

    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == pangkalan_id:
            # Update existing row
            ws.cell(row=row, column=date_col_start, value=stok_int)
            ws.cell(row=row, column=date_col_start + 1, value=inputan_int)
            ws.cell(row=row, column=date_col_start + 2, value=timestamp)
            pangkalan_exists = True
            break

    if not pangkalan_exists:
        # Tambahkan row baru
        new_row = data_start_row
        ws.cell(row=new_row, column=1, value=pangkalan_id)
        ws.cell(row=new_row, column=2, value=nama_pangkalan)
        ws.cell(row=new_row, column=date_col_start, value=stok_int)
        ws.cell(row=new_row, column=date_col_start + 1, value=inputan_int)
        ws.cell(row=new_row, column=date_col_start + 2, value=timestamp)


def _apply_conditional_formatting(
    ws, stok_int, inputan_int, pangkalan_id, date_col_start
):
    """Apply conditional formatting untuk stok > 90 dan input = 0"""
    if stok_int is not None and inputan_int is not None:
        if stok_int > 90 and inputan_int == 0:
            # Cari row yang tepat untuk conditional formatting
            for row in range(3, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == pangkalan_id:
                    yellow_fill = PatternFill(
                        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                    )
                    ws.cell(row=row, column=date_col_start).fill = yellow_fill  # STOK
                    ws.cell(
                        row=row, column=date_col_start + 1
                    ).fill = yellow_fill  # INPUT
                    print(
                        f"Applied yellow highlight: STOK={stok_int} (>90) && INPUT={inputan_int} (=0)"
                    )
                    break


def _apply_conditional_formatting_new(
    ws, stok_int, inputan_int, pangkalan_row, date_col_start
):
    """Apply conditional formatting - hanya kuning untuk STOK > 90 && INPUT = 0"""
    if stok_int is not None and inputan_int is not None:
        # Hanya satu kondisi: STOK > 90 dan INPUT = 0 (Kuning)
        if stok_int > 90 and inputan_int == 0:
            yellow_fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )
            ws.cell(row=pangkalan_row, column=date_col_start).fill = yellow_fill  # STOK
            ws.cell(
                row=pangkalan_row, column=date_col_start + 1
            ).fill = yellow_fill  # INPUT
            print(
                f"Applied yellow highlight: STOK={stok_int} (>90) && INPUT={inputan_int} (=0)"
            )


def _format_headers_new(ws, display_date):
    """Format headers baru untuk sistem incremental"""
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Apply center alignment untuk semua headers
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).alignment = center_alignment
        ws.cell(row=2, column=col).alignment = center_alignment

    # Merge cells untuk header tanggal yang baru ditambahkan
    _merge_and_color_date_headers(ws, display_date, center_alignment)


def _format_headers(ws, display_date):
    """Format headers dengan merge cell dan center alignment dengan fill color berdasarkan hari"""
    # Hanya center alignment untuk headers, tanpa bold atau background color
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Apply center alignment to row 1 dan 2 tanpa font bold atau background color
    for col in range(1, ws.max_column + 1):
        # Row 1 - hanya center alignment
        ws.cell(row=1, column=col).alignment = center_alignment
        # Row 2 - hanya center alignment
        ws.cell(row=2, column=col).alignment = center_alignment

    # Merge cells untuk header tanggal dan apply colors
    _merge_and_color_date_headers(ws, display_date, center_alignment)


def _merge_and_color_date_headers(ws, display_date, center_alignment):
    """Merge cells untuk header tanggal dan apply colors berdasarkan hari"""
    # Cari semua kolom yang memiliki header tanggal (format YYYY-MM-DD)
    date_start_cols = []

    for col in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=1, column=col).value
        if header_cell:
            # Deteksi apakah ini header tanggal dengan pattern YYYY-MM-DD
            is_date_header = re.match(r"^\d{4}-\d{2}-\d{2}$", str(header_cell))
            is_not_basic_header = str(header_cell) not in [
                "PANGKALAN_ID",
                "NAMA_PANGKALAN",
            ]

            if is_date_header or (is_not_basic_header and col > 2):
                # Cek apakah ini kolom pertama dari grup (harus diikuti STOK, INPUT, TIME, STATUS di row 2)
                if col + 3 <= ws.max_column:
                    row2_col1 = ws.cell(row=2, column=col).value
                    row2_col2 = ws.cell(row=2, column=col + 1).value
                    row2_col3 = ws.cell(row=2, column=col + 2).value
                    row2_col4 = ws.cell(row=2, column=col + 3).value

                    if (
                        row2_col1 == "STOK (TABUNG)"
                        and row2_col2 == "INPUT (TABUNG)"
                        and row2_col3 == "TIME"
                        and row2_col4 == "STATUS"
                    ):
                        date_start_cols.append(col)

    # Apply merge dan color untuk setiap grup tanggal
    for start_col in date_start_cols:
        end_col = min(
            start_col + 3, ws.max_column
        )  # 4 kolom: STOK, INPUT, TIME, STATUS
        if end_col > start_col:
            try:
                # Merge cell di row 1 untuk header tanggal
                ws.merge_cells(
                    start_row=1, start_column=start_col, end_row=1, end_column=end_col
                )

                # Clear isi di kolom yang di-merge (kecuali kolom pertama)
                for clear_col in range(start_col + 1, end_col + 1):
                    ws.cell(row=1, column=clear_col).value = None

                # Set alignment untuk cell yang di-merge (tanpa warna)
                merged_cell = ws.cell(row=1, column=start_col)
                merged_cell.alignment = center_alignment

                print(f"✓ Merged header untuk kolom {start_col}-{end_col}")

            except Exception as e:
                print(
                    f"⚠ Warning: Failed to merge cells for date group {start_col}-{end_col}: {e}"
                )


def _apply_borders(ws):
    """Apply border untuk semua cell"""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Terapkan border ke semua cell yang memiliki data
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

    # Apply conditional formatting untuk semua data yang ada
    _apply_all_conditional_formatting(ws)


def _apply_all_conditional_formatting(ws):
    """Apply conditional formatting dan center alignment untuk semua data yang ada"""
    max_row = ws.max_row
    max_col = ws.max_column
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    print("Checking for conditional formatting conditions...")

    # Cari semua kolom yang berisi data stok dan input
    for col in range(1, max_col + 1):
        # Cek apakah ini kolom STOK (TABUNG) atau INPUT (TABUNG)
        row2_header = ws.cell(row=2, column=col).value
        if row2_header == "STOK (TABUNG)":
            # Ini kolom stok, cari kolom input yang bersebelahan (biasanya +1)
            input_col = col + 1
            time_col = col + 2  # Kolom TIME
            status_col = col + 3  # Kolom STATUS

            # Loop melalui semua data rows (mulai dari row 3)
            for row in range(3, max_row + 1):
                try:
                    # Apply center alignment untuk semua data (STOK, INPUT, TIME, STATUS)
                    if col <= max_col:
                        ws.cell(
                            row=row, column=col
                        ).alignment = center_alignment  # STOK
                    if (
                        input_col <= max_col
                        and ws.cell(row=2, column=input_col).value == "INPUT (TABUNG)"
                    ):
                        ws.cell(
                            row=row, column=input_col
                        ).alignment = center_alignment  # INPUT
                    if time_col <= max_col:
                        ws.cell(
                            row=row, column=time_col
                        ).alignment = center_alignment  # TIME
                    if (
                        status_col <= max_col
                        and ws.cell(row=2, column=status_col).value == "STATUS"
                    ):
                        ws.cell(
                            row=row, column=status_col
                        ).alignment = center_alignment  # STATUS

                    # Conditional formatting untuk STOK dan INPUT
                    if (
                        input_col <= max_col
                        and ws.cell(row=2, column=input_col).value == "INPUT (TABUNG)"
                    ):
                        stok_value = ws.cell(row=row, column=col).value
                        input_value = ws.cell(row=row, column=input_col).value

                        # Cek kondisi: stok > 90 dan input = 0
                        if stok_value is not None and input_value is not None:
                            stok_num = (
                                int(stok_value)
                                if isinstance(stok_value, (int, str))
                                and str(stok_value).isdigit()
                                else None
                            )
                            input_num = (
                                int(input_value)
                                if isinstance(input_value, (int, str))
                                and str(input_value).isdigit()
                                else None
                            )

                            if stok_num is not None and input_num is not None:
                                if stok_num > 90 and input_num == 0:
                                    # Apply warna kuning
                                    ws.cell(
                                        row=row, column=col
                                    ).fill = yellow_fill  # STOK
                                    ws.cell(
                                        row=row, column=input_col
                                    ).fill = yellow_fill  # INPUT
                                    print(
                                        f"Applied yellow highlight at row {row}: STOK={stok_num} (>90) && INPUT={input_num} (=0)"
                                    )
                except (ValueError, TypeError):
                    # Skip jika tidak bisa convert ke int
                    continue


def save_to_excel_new_format(
    nama_pangkalan,
    tanggal_check,
    stok_awal,
    total_inputan,
    status,
    selected_date=None,
    pangkalan_id=None,
):
    """Simpan data ke Excel dengan format baru (backward compatibility)"""
    import pandas as pd

    filename = get_excel_filename(selected_date)
    filepath = os.path.join(RESULTS_DIR, filename)

    try:
        print(f"Menyimpan data ke Excel ({filename})...")

        # Siapkan data dengan format baru
        data = {
            "NAMA PANGKALAN": [nama_pangkalan],
            "TANGGAL CHECK": [tanggal_check],
            "STOK AWAL": [stok_awal or "Tidak Ditemukan"],
            "TOTAL INPUTAN": [total_inputan or "Tidak Ditemukan"],
            "STATUS": [status],
        }
        if pangkalan_id:
            data["PANGKALAN_ID"] = [pangkalan_id]

        df_new = pd.DataFrame(data)

        if os.path.exists(filepath):
            df_existing = pd.read_excel(filepath)

            # Tambahkan data baru ke DataFrame yang sudah ada
            df_existing = pd.concat([df_existing, df_new], ignore_index=True)

            df_existing.to_excel(filepath, index=False)
            print(f"✓ Data berhasil disimpan ke: {filepath}")
            print("\nHasil scraping:")
            print(df_new.to_string(index=False))
        else:
            df_new.to_excel(filepath, index=False)
            print(f"✓ Data berhasil disimpan ke: {filepath}")
            print("\nHasil scraping:")
            print(df_new.to_string(index=False))
    except Exception as e:
        print(f"✗ Error saat menyimpan: {str(e)}")
        import logging

        logger = logging.getLogger("automation")
        logger.error(f"Error saving to Excel: {str(e)}", exc_info=True)
